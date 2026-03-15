from flask import Flask, render_template, request, redirect, url_for, session, flash
import psycopg2, psycopg2.extras, psycopg2.pool, random, string, hashlib, os, json, click
from datetime import datetime, timezone, timedelta
from functools import wraps

# Load .env file automatically when running locally
# (python-dotenv is optional — skipped silently if not installed)
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

# Kenya is UTC+3 (East Africa Time) — no DST observed
EAT = timezone(timedelta(hours=3))

def now_eat():
    """Return the current moment as a naive datetime in East Africa Time (UTC+3)."""
    return datetime.now(EAT).replace(tzinfo=None)

def coerce_dt(val):
    """Accept either a Python datetime (from psycopg2) or a string and return a naive datetime."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.replace(tzinfo=None)
    return datetime.strptime(str(val)[:19], '%Y-%m-%d %H:%M:%S')

app = Flask(__name__)

_secret = os.environ.get('SECRET_KEY')
if not _secret:
    import warnings
    warnings.warn(
        "SECRET_KEY env var is not set. Using a random key — all sessions will be "
        "lost on restart. Set SECRET_KEY=<random 32-char string> in production.",
        stacklevel=2
    )
    _secret = os.urandom(32)
app.secret_key = _secret
ADMIN_PASSWORD_INIT = os.environ.get('ADMIN_PASSWORD', 'changeme')

@app.context_processor
def inject_globals():
    return dict(json=json)

@app.template_filter('dt_fmt')
def dt_fmt(value, fmt='%Y-%m-%d'):
    """Format a datetime column (object or string) with a strftime pattern.
    Default fmt='%Y-%m-%d' gives YYYY-MM-DD (replaces the old [:10] slicing).
    Pass fmt='%Y-%m-%d %H:%M' for YYYY-MM-DD HH:MM (replaces [:16] slicing).
    Returns '—' for None/empty values.
    """
    if not value:
        return '—'
    try:
        if isinstance(value, str):
            value = datetime.strptime(value[:19], '%Y-%m-%d %H:%M:%S')
        return value.strftime(fmt)
    except Exception:
        return str(value)

@app.template_filter('eat_fmt')
def eat_fmt(value, fmt='%d %b %Y, %I:%M %p'):
    """Format a stored EAT datetime string for display. e.g. '14 Jun 2025, 03:45 PM EAT'"""
    if not value:
        return '—'
    try:
        if isinstance(value, str):
            value = datetime.strptime(value[:19], '%Y-%m-%d %H:%M:%S')
        return value.strftime(fmt) + ' EAT'
    except Exception:
        return str(value)

# ─── DB connection pool ───────────────────────────────────────────────────────
# On cPanel shared hosting, Python apps run under Phusion Passenger (not
# Waitress).  Passenger forks worker processes, so the pool is created lazily
# per-process to avoid sharing a pool across forks.
#
# Pool sizing:
#   minconn=1  – keep 1 connection warm per worker process
#   maxconn=5  – cap at 5 per worker; with Passenger's typical 2-4 workers
#                that stays well under the shared-host pg max_connections (25)
#
# If you ever move to a VPS and raise pg max_connections you can increase these.

_pool = None

def _get_pool():
    global _pool
    if _pool is None:
        _pool = psycopg2.pool.ThreadedConnectionPool(
            minconn=1,
            maxconn=5,
            host=os.environ.get('DB_HOST', 'localhost'),
            port=int(os.environ.get('DB_PORT', 5432)),
            dbname=os.environ.get('DB_NAME', 'bible_trivia'),
            user=os.environ.get('DB_USER', 'bible_trivia_user'),
            password=os.environ.get('DB_PASSWORD', ''),
            connect_timeout=10,
        )
    return _pool

def get_db():
    """Borrow a connection from the pool.
    Always pair with close_db(conn) — even in error paths — so the
    connection is returned and not leaked.
    Set DB credentials in cPanel > Software > Setup Python App > Environment Variables:
      DB_HOST, DB_PORT, DB_NAME, DB_USER, DB_PASSWORD
    """
    conn = _get_pool().getconn()
    conn.autocommit = False
    # Use DictCursor so columns are accessible by name (like sqlite3.Row)
    conn.cursor_factory = psycopg2.extras.RealDictCursor
    return conn

def close_db(conn):
    """Return a connection back to the pool (not actually closed)."""
    if conn is not None:
        try:
            # Roll back any uncommitted transaction so the connection is
            # clean before being reused by the next request.
            conn.rollback()
        except Exception:
            pass
        try:
            _get_pool().putconn(conn)
        except Exception:
            pass

def _exec(conn, sql, params=()):
    """Execute a statement, return cursor."""
    cur = conn.cursor()
    cur.execute(sql, params)
    return cur

def _fetchone(conn, sql, params=()):
    """Execute and return one row as a dict-like object."""
    cur = conn.cursor()
    cur.execute(sql, params)
    row = cur.fetchone()
    cur.close()
    return row

def _fetchall(conn, sql, params=()):
    """Execute and return all rows."""
    cur = conn.cursor()
    cur.execute(sql, params)
    rows = cur.fetchall()
    cur.close()
    return rows

def _lastrowid(conn, sql, params=()):
    """Execute an INSERT and return the new row id via RETURNING id."""
    cur = conn.cursor()
    # Append RETURNING id if not already present
    sql_r = sql.rstrip().rstrip(';')
    if 'RETURNING' not in sql_r.upper():
        sql_r += ' RETURNING id'
    cur.execute(sql_r, params)
    row = cur.fetchone()
    cur.close()
    return row['id'] if row else None


# ─── Audit logging ────────────────────────────────────────────────────────────

def log_action(conn, action, category='admin', entity_type=None,
               entity_id=None, entity_name=None, details=None):
    """Write an audit log entry into the open connection.
    Call BEFORE conn.commit() so the log is atomic with the main operation.
    Never raises — failures are silently swallowed so a logging error can never
    break a real operation.

    Args:
        conn        : open psycopg2 connection
        action      : short snake_case identifier  e.g. 'create_session'
        category    : 'admin' | 'user' | 'system'
        entity_type : 'session' | 'section' | 'question' | 'user' | 'audit_log' …
        entity_id   : integer PK of the affected row (optional)
        entity_name : human-readable label  e.g. session name (optional)
        details     : free-text note or JSON snippet (optional)
    """
    try:
        ip = None
        try:
            ip = request.remote_addr
        except RuntimeError:
            pass  # called outside request context (e.g. CLI)
        _exec(conn, '''
            INSERT INTO audit_logs
                (action, category, entity_type, entity_id, entity_name, details, ip_address)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        ''', (action, category, entity_type, entity_id, entity_name, details, ip))
    except Exception:
        pass  # never crash the caller


def init_db():
    conn = get_db()
    cur = conn.cursor()
    tables = [
        """CREATE TABLE IF NOT EXISTS users (
            id         SERIAL PRIMARY KEY,
            phone      TEXT UNIQUE NOT NULL,
            name       TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT (NOW() AT TIME ZONE 'Africa/Nairobi')
        )""",
        """CREATE TABLE IF NOT EXISTS quiz_sessions (
            id                  SERIAL PRIMARY KEY,
            name                TEXT NOT NULL,
            description         TEXT DEFAULT '',
            is_active           INTEGER DEFAULT 1,
            randomize_questions INTEGER DEFAULT 1,
            time_limit_minutes  INTEGER DEFAULT 0,
            scheduled_start     TIMESTAMP DEFAULT NULL,
            created_at          TIMESTAMP DEFAULT (NOW() AT TIME ZONE 'Africa/Nairobi')
        )""",
        """CREATE TABLE IF NOT EXISTS sections (
            id         SERIAL PRIMARY KEY,
            session_id INTEGER NOT NULL REFERENCES quiz_sessions(id) ON DELETE CASCADE,
            name       TEXT NOT NULL,
            order_num  INTEGER DEFAULT 0
        )""",
        """CREATE TABLE IF NOT EXISTS questions (
            id             SERIAL PRIMARY KEY,
            section_id     INTEGER NOT NULL REFERENCES sections(id) ON DELETE CASCADE,
            question_type  TEXT DEFAULT 'single',
            question_text  TEXT NOT NULL,
            option_a       TEXT DEFAULT '',
            option_b       TEXT DEFAULT '',
            option_c       TEXT DEFAULT '',
            option_d       TEXT DEFAULT '',
            correct_answer TEXT NOT NULL,
            blank_options  TEXT DEFAULT '[]',
            points         INTEGER DEFAULT 1,
            order_num      INTEGER DEFAULT 0
        )""",
        """CREATE TABLE IF NOT EXISTS user_sessions (
            id           SERIAL PRIMARY KEY,
            user_id      INTEGER NOT NULL REFERENCES users(id),
            session_id   INTEGER NOT NULL REFERENCES quiz_sessions(id),
            started_at   TIMESTAMP DEFAULT (NOW() AT TIME ZONE 'Africa/Nairobi'),
            completed_at TIMESTAMP
        )""",
        """CREATE TABLE IF NOT EXISTS user_answers (
            id               SERIAL PRIMARY KEY,
            user_session_id  INTEGER NOT NULL REFERENCES user_sessions(id),
            question_id      INTEGER NOT NULL REFERENCES questions(id),
            selected_answer  TEXT NOT NULL,
            is_correct       INTEGER NOT NULL,
            reward_code      TEXT,
            answered_at      TIMESTAMP DEFAULT (NOW() AT TIME ZONE 'Africa/Nairobi')
        )""",
        """CREATE TABLE IF NOT EXISTS app_settings (
            key   TEXT PRIMARY KEY,
            value TEXT
        )""",
        """CREATE TABLE IF NOT EXISTS cheat_flags (
            id              SERIAL PRIMARY KEY,
            user_session_id INTEGER NOT NULL REFERENCES user_sessions(id) ON DELETE CASCADE,
            violation_type  TEXT NOT NULL,
            flagged_at      TIMESTAMP DEFAULT (NOW() AT TIME ZONE 'Africa/Nairobi')
        )""",
        """CREATE TABLE IF NOT EXISTS audit_logs (
            id          SERIAL PRIMARY KEY,
            action      TEXT NOT NULL,
            category    TEXT NOT NULL DEFAULT 'admin',
            entity_type TEXT,
            entity_id   INTEGER,
            entity_name TEXT,
            details     TEXT,
            ip_address  TEXT,
            logged_at   TIMESTAMP DEFAULT (NOW() AT TIME ZONE 'Africa/Nairobi')
        )""",
    ]
    for ddl in tables:
        cur.execute(ddl)
    cur.execute(
        "INSERT INTO app_settings (key, value) VALUES (%s, %s) ON CONFLICT (key) DO NOTHING",
        ('admin_password', ADMIN_PASSWORD_INIT)
    )
    conn.commit()
    # Migrate: add columns that may not exist yet (PostgreSQL IF NOT EXISTS)
    migrations = [
        "ALTER TABLE questions ADD COLUMN IF NOT EXISTS question_type TEXT DEFAULT 'single'",
        "ALTER TABLE questions ADD COLUMN IF NOT EXISTS blank_options TEXT DEFAULT '[]'",
        "ALTER TABLE quiz_sessions ADD COLUMN IF NOT EXISTS time_limit_minutes INTEGER DEFAULT 0",
        "ALTER TABLE quiz_sessions ADD COLUMN IF NOT EXISTS scheduled_start TIMESTAMP DEFAULT NULL",
        "ALTER TABLE user_answers ADD COLUMN IF NOT EXISTS points_earned NUMERIC(8,2) DEFAULT 0",
        # audit_logs — create if it doesn't exist yet (for existing deployments)
        """CREATE TABLE IF NOT EXISTS audit_logs (
            id          SERIAL PRIMARY KEY,
            action      TEXT NOT NULL,
            category    TEXT NOT NULL DEFAULT 'admin',
            entity_type TEXT,
            entity_id   INTEGER,
            entity_name TEXT,
            details     TEXT,
            ip_address  TEXT,
            logged_at   TIMESTAMP DEFAULT (NOW() AT TIME ZONE 'Africa/Nairobi')
        )""",
    ]
    for sql in migrations:
        try:
            cur.execute(sql)
            conn.commit()
        except Exception:
            conn.rollback()
    cur.close()
    close_db(conn)

def normalize_multi(answer_str):
    """Sort comma-separated letters for comparison: 'C,A' -> 'A,C'"""
    return ','.join(sorted(x.strip().upper() for x in answer_str.split(',') if x.strip()))

def score_answer(question, selected_raw):
    """
    Returns (is_correct: int, stored_selected: str, points_earned: float).

    Scoring rules:
      single     – full marks or 0 (unchanged).
      multi      – marks split evenly across correct options.
                   Selecting any WRONG option forfeits all points.
                   Selecting a correct subset (no wrong options) earns
                   proportional marks.  Full marks + is_correct=1 only
                   when the user selects every correct option and nothing else.
      fill_blank – marks split evenly across blanks.
                   Each correctly-filled blank earns its share of the points.
                   is_correct=1 only when every blank is right.
    """
    qtype   = question['question_type'] or 'single'
    correct = (question['correct_answer'] or '').strip()
    points  = float(question.get('points') or 1)

    if qtype == 'single':
        sel = selected_raw.strip().upper()
        is_correct = int(sel == correct.upper())
        return is_correct, sel, points if is_correct else 0.0

    elif qtype == 'multi':
        sel_norm  = normalize_multi(selected_raw)
        corr_norm = normalize_multi(correct)
        sel_set   = set(sel_norm.split(',')) if sel_norm  else set()
        corr_set  = set(corr_norm.split(',')) if corr_norm else set()

        if not corr_set:
            return 0, sel_norm, 0.0

        # Credit every correctly-selected option regardless of whether
        # wrong options were also chosen. Points split evenly across
        # all correct options.
        matched        = sel_set & corr_set
        pts_per_option = points / len(corr_set)
        earned         = round(len(matched) * pts_per_option, 2)
        # is_correct=1 only when the selection matches exactly
        is_correct     = int(sel_norm == corr_norm)
        return is_correct, sel_norm, earned

    elif qtype == 'fill_blank':
        sel_parts  = [p.strip() for p in selected_raw.split('|')]
        corr_parts = [p.strip() for p in correct.split('|')]
        if not corr_parts:
            return 0, selected_raw, 0.0

        pts_per_blank = points / len(corr_parts)
        correct_count = sum(1 for s, c in zip(sel_parts, corr_parts) if s == c)
        earned        = round(correct_count * pts_per_blank, 2)
        is_correct    = int(sel_parts == corr_parts)
        return is_correct, selected_raw, earned

    return 0, selected_raw, 0.0

# Keep old name as alias so any external callers still work
def check_answer(question, selected_raw):
    is_correct, stored, _ = score_answer(question, selected_raw)
    return is_correct, stored

def get_remaining_seconds(user_session_row, time_limit_minutes):
    """Return seconds left (None = no limit, 0 = expired). Uses EAT throughout."""
    if not time_limit_minutes:
        return None
    started = coerce_dt(user_session_row['started_at'])
    elapsed = (now_eat() - started).total_seconds()
    remaining = int(time_limit_minutes * 60 - elapsed)
    return max(remaining, 0)

def parse_scheduled_start(raw):
    """Convert datetime-local input (YYYY-MM-DDTHH:MM) to DB format (YYYY-MM-DD HH:MM:SS), or None."""
    if not raw or not raw.strip():
        return None
    raw = raw.strip().replace('T', ' ')
    if len(raw) == 16:
        raw = raw + ':00'
    try:
        datetime.strptime(raw, '%Y-%m-%d %H:%M:%S')
        return raw
    except ValueError:
        return None

def generate_code(user_id, question_id):
    raw = f"{user_id}-{question_id}-{random.randint(10000,99999)}"
    return hashlib.md5(raw.encode()).hexdigest()[:8].upper()

# ─── Auth decorators ──────────────────────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('is_admin'):
            return redirect(url_for('admin_login'))
        return f(*args, **kwargs)
    return decorated

def normalize_phone(raw):
    """Normalize any Kenyan phone format to 07XXXXXXXX or 01XXXXXXXX (10 digits)."""
    phone = raw.strip().replace(' ', '').replace('-', '')
    # +2547... or 2547... → 07...
    if phone.startswith('+254'):
        phone = '0' + phone[4:]
    elif phone.startswith('254') and len(phone) >= 12:
        phone = '0' + phone[3:]
    return phone

# ═══════════════════════════════════════════════════════════════════════════════
#  USER ROUTES
# ═══════════════════════════════════════════════════════════════════════════════

@app.route('/', methods=['GET', 'POST'])
def index():
    if 'user_id' in session:
        return redirect(url_for('quiz_home'))
    if request.method == 'POST':
        phone = normalize_phone(request.form.get('phone', ''))
        if not phone:
            flash('Please enter your phone number.', 'error')
            return render_template('index.html')
        conn = get_db()
        user = _fetchone(conn, 'SELECT * FROM users WHERE phone=%s', (phone,))
        if user:
            session['user_id']   = user['id']
            session['user_name'] = user['name']
            log_action(conn, 'user_login', category='user',
                       entity_type='user', entity_id=user['id'], entity_name=user['name'],
                       details=f"{user['name']} logged in ({phone})")
            conn.commit()
            close_db(conn)
            return redirect(url_for('quiz_home'))
        close_db(conn)
        session['pending_phone'] = phone
        return redirect(url_for('register'))
    return render_template('index.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if 'pending_phone' not in session:
        return redirect(url_for('index'))
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        if not name:
            flash('Please enter your name.', 'error')
            return render_template('register.html', phone=session['pending_phone'])
        phone = session.pop('pending_phone')
        conn = get_db()
        _exec(conn, 'INSERT INTO users (phone, name) VALUES (%s,%s) ON CONFLICT (phone) DO NOTHING', (phone, name))
        log_action(conn, 'user_register', category='user',
                   entity_type='user', entity_name=name,
                   details=f"New user registered: {name} ({phone})")
        conn.commit()
        user = _fetchone(conn, 'SELECT * FROM users WHERE phone=%s', (phone,))
        close_db(conn)
        session['user_id']   = user['id']
        session['user_name'] = user['name']
        return redirect(url_for('quiz_home'))
    return render_template('register.html', phone=session['pending_phone'])

@app.route('/quiz')
@login_required
def quiz_home():
    conn = get_db()
    real_user = _fetchone(conn, 'SELECT id FROM users WHERE id=%s', (session['user_id'],))
    if not real_user:
        close_db(conn)
        session.clear()
        flash('Your session has expired. Please log in again.', 'error')
        return redirect(url_for('index'))
    sessions_list = _fetchall(conn,
        'SELECT * FROM quiz_sessions WHERE is_active=1 ORDER BY created_at DESC'
    )
    completed_ids = {
        r['session_id'] for r in
        _fetchall(conn, 'SELECT session_id FROM user_sessions WHERE user_id=%s AND completed_at IS NOT NULL',
                     (session['user_id'],))
    }
    # in-progress: include started_at so we can show live countdown in the modal
    inprogress_rows = _fetchall(conn,
        'SELECT session_id, started_at FROM user_sessions WHERE user_id=%s AND completed_at IS NULL',
        (session['user_id'],)
    )
    inprogress_ids = {r['session_id'] for r in inprogress_rows}
    # Map session_id -> remaining seconds (None if no limit)
    inprogress_remaining = {}
    for row in inprogress_rows:
        sid = row['session_id']
        qs_row = next((s for s in sessions_list if s['id'] == sid), None)
        if qs_row:
            rem = get_remaining_seconds(row, qs_row['time_limit_minutes'] or 0)
            inprogress_remaining[sid] = rem  # None = no limit, int = seconds left
    close_db(conn)
    # Build scheduled_start map for frontend (seconds until start, or 0 if past)
    now = now_eat()
    scheduled_info = {}
    _DAYS   = ['Mon','Tue','Wed','Thu','Fri','Sat','Sun']
    _MONTHS = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
    for s in sessions_list:
        if s['scheduled_start']:
            sched = coerce_dt(s['scheduled_start'])
            diff  = int((sched - now).total_seconds())
            # Human readable: "Sat 14 Mar 2026, 1:18 PM"
            hour   = sched.hour % 12 or 12
            ampm   = 'AM' if sched.hour < 12 else 'PM'
            pretty = (f"{_DAYS[sched.weekday()]} {sched.day} {_MONTHS[sched.month-1]} "
                      f"{sched.year}, {hour}:{sched.minute:02d} {ampm}")
            scheduled_info[s['id']] = {
                'sched_str':    s['scheduled_start'],   # raw, for DB comparison
                'pretty':       pretty,                  # display string
                'seconds_until': max(diff, 0),
                'started':      diff <= 0,
            }
        else:
            scheduled_info[s['id']] = None
    return render_template('quiz_home.html', sessions=sessions_list,
                           completed_ids=completed_ids, inprogress_ids=inprogress_ids,
                           inprogress_remaining=inprogress_remaining,
                           scheduled_info=scheduled_info)

@app.route('/quiz/<int:session_id>/start', methods=['POST'])
@login_required
def start_quiz(session_id):
    """Called when the user explicitly clicks 'Let's Go' — creates the user_session record (starts the timer)."""
    conn = get_db()
    real_user = _fetchone(conn, 'SELECT id FROM users WHERE id=%s', (session['user_id'],))
    if not real_user:
        close_db(conn); session.clear()
        flash('Your session has expired. Please log in again.', 'error')
        return redirect(url_for('index'))

    qs = _fetchone(conn, 'SELECT * FROM quiz_sessions WHERE id=%s AND is_active=1', (session_id,))
    if not qs:
        flash('Session not found or inactive.', 'error')
        close_db(conn)
        return redirect(url_for('quiz_home'))

    # Check scheduled start
    if qs['scheduled_start']:
        sched = coerce_dt(qs['scheduled_start'])
        if now_eat() < sched:
            flash('This session has not started yet. Please wait until the scheduled time.', 'error')
            close_db(conn)
            return redirect(url_for('quiz_home'))

    # Check if already in progress — just resume
    us = _fetchone(conn,
        'SELECT * FROM user_sessions WHERE user_id=%s AND session_id=%s AND completed_at IS NULL',
        (session['user_id'], session_id)
    )
    if not us:
        _exec(conn, 'INSERT INTO user_sessions (user_id, session_id) VALUES (%s,%s)',
                     (session['user_id'], session_id))
        log_action(conn, 'quiz_start', category='user',
                   entity_type='session', entity_id=session_id, entity_name=qs['name'],
                   details=f"{session.get('user_name')} started quiz '{qs['name']}'")
        conn.commit()
    close_db(conn)
    return redirect(url_for('take_quiz', session_id=session_id))

@app.route('/quiz/<int:session_id>', methods=['GET', 'POST'])
@login_required
def take_quiz(session_id):
    conn = get_db()

    # ── Guard: verify session cookie user still exists in DB ──────────────
    # Happens when DB is wiped but browser still holds the old session cookie
    real_user = _fetchone(conn, 'SELECT id FROM users WHERE id=%s', (session['user_id'],))
    if not real_user:
        close_db(conn)
        session.clear()
        flash('Your session has expired. Please log in again.', 'error')
        return redirect(url_for('index'))

    qs = _fetchone(conn, 'SELECT * FROM quiz_sessions WHERE id=%s AND is_active=1', (session_id,))
    if not qs:
        flash('Session not found or is inactive.', 'error')
        close_db(conn)
        return redirect(url_for('quiz_home'))

    # Get existing user_session — do NOT create one here (that's done in start_quiz)
    us = _fetchone(conn,
        'SELECT * FROM user_sessions WHERE user_id=%s AND session_id=%s AND completed_at IS NULL',
        (session['user_id'], session_id)
    )
    if not us:
        # No active session — user hasn't started yet or already completed
        completed = _fetchone(conn,
            'SELECT id FROM user_sessions WHERE user_id=%s AND session_id=%s AND completed_at IS NOT NULL',
            (session['user_id'], session_id)
        )
        close_db(conn)
        if completed:
            return redirect(url_for('results', session_id=session_id))
        # Haven't started — send back to home to click Start
        flash('Please click "Start Quiz" to begin.', 'error')
        return redirect(url_for('quiz_home'))

    us_id = us['id']

    # ── Timer check ────────────────────────────────────────────────────────
    time_limit = qs['time_limit_minutes'] or 0
    remaining_seconds = get_remaining_seconds(us, time_limit)
    if remaining_seconds is not None and remaining_seconds <= 0:
        # Time is up — auto-complete the session
        _exec(conn, "UPDATE user_sessions SET completed_at=(NOW() AT TIME ZONE 'Africa/Nairobi') WHERE id=%s", (us_id,))
        conn.commit()
        close_db(conn)
        flash('⏰ Time is up! Your session has been submitted.', 'error')
        return redirect(url_for('results', session_id=session_id))
    sections = _fetchall(conn,
        'SELECT * FROM sections WHERE session_id=%s ORDER BY order_num', (session_id,)
    )
    all_questions = []
    for sec in sections:
        qs_list = _fetchall(conn,
            'SELECT q.*, %s::text as section_name FROM questions q WHERE q.section_id=%s ORDER BY q.order_num',
            (sec['name'], sec['id'])
        )
        all_questions.extend(qs_list)

    # Randomize per user_session (stable seed so page reloads keep same order)
    if qs['randomize_questions']:
        r = random.Random(us_id)
        r.shuffle(all_questions)

    answered = _fetchall(conn,
        'SELECT * FROM user_answers WHERE user_session_id=%s', (us_id,)
    )
    answered_map = {a['question_id']: a for a in answered}
    answered_ids = set(answered_map.keys())

    if request.method == 'POST':
        q_id = int(request.form.get('question_id'))
        if q_id not in answered_ids:
            question = _fetchone(conn, 'SELECT * FROM questions WHERE id=%s', (q_id,))
            qtype = question['question_type'] or 'single'

            if qtype == 'single':
                selected_raw = request.form.get('answer', '').strip().upper()
            elif qtype == 'multi':
                checked = request.form.getlist('answer')
                selected_raw = ','.join(sorted(x.upper() for x in checked)) if checked else ''
            elif qtype == 'fill_blank':
                bo = json.loads(question['blank_options'] or '[]')
                parts = [request.form.get(f'blank_{i}', '').strip() for i in range(len(bo))]
                selected_raw = '|'.join(parts)
            else:
                selected_raw = ''

            if selected_raw:
                is_correct, stored_sel, pts_earned = score_answer(question, selected_raw)
                code = generate_code(session['user_id'], q_id) if is_correct else None
                _exec(conn,
                    'INSERT INTO user_answers (user_session_id, question_id, selected_answer, is_correct, points_earned, reward_code) VALUES (%s,%s,%s,%s,%s,%s)',
                    (us_id, q_id, stored_sel, is_correct, pts_earned, code)
                )
                result_label = 'correct' if is_correct else 'wrong'
                log_action(conn, 'quiz_answer', category='user',
                           entity_type='question', entity_id=q_id,
                           entity_name=session.get('user_name'),
                           details=f"Q#{q_id} answered {result_label} in session #{session_id}")
                conn.commit()
                answered_ids.add(q_id)

        if len(answered_ids) >= len(all_questions):
            _exec(conn, "UPDATE user_sessions SET completed_at=(NOW() AT TIME ZONE 'Africa/Nairobi') WHERE id=%s", (us_id,))
            log_action(conn, 'quiz_complete', category='user',
                       entity_type='session', entity_id=session_id, entity_name=qs['name'],
                       details=f"{session.get('user_name')} completed '{qs['name']}' "
                               f"({len(answered_ids)}/{len(all_questions)} answered)")
            conn.commit()
            close_db(conn)
            return redirect(url_for('results', session_id=session_id))
        close_db(conn)
        return redirect(url_for('take_quiz', session_id=session_id))

    # Find next unanswered
    next_q = next((q for q in all_questions if q['id'] not in answered_ids), None)
    if not next_q:
        _exec(conn, "UPDATE user_sessions SET completed_at=(NOW() AT TIME ZONE 'Africa/Nairobi') WHERE id=%s", (us_id,))
        conn.commit()
        close_db(conn)
        return redirect(url_for('results', session_id=session_id))

    # Existing cheat flag count — needed by the anti-cheat JS to restore strike state
    existing_flags_row = _fetchone(conn,
        'SELECT COUNT(*) as n FROM cheat_flags WHERE user_session_id=%s', (us_id,)
    )
    existing_flags = int(existing_flags_row['n']) if existing_flags_row else 0

    close_db(conn)
    return render_template('quiz.html', question=next_q, quiz_session=qs,
                           progress=len(answered_ids), total=len(all_questions),
                           all_questions=all_questions, answered_map=answered_map,
                           answered_ids=answered_ids,
                           remaining_seconds=remaining_seconds,
                           time_limit=time_limit,
                           existing_flags=existing_flags,
                           quiz_mode=True)

@app.route('/quiz/<int:session_id>/expire', methods=['POST'])
@login_required
def expire_quiz(session_id):
    conn = get_db()
    qs_row = _fetchone(conn, 'SELECT name FROM quiz_sessions WHERE id=%s', (session_id,))
    _exec(conn, "UPDATE user_sessions SET completed_at=(NOW() AT TIME ZONE 'Africa/Nairobi') WHERE user_id=%s AND session_id=%s AND completed_at IS NULL",
                 (session['user_id'], session_id))
    reason = request.form.get('reason', '')
    action_label = 'quiz_auto_submit_cheat' if reason == 'cheat' else 'quiz_time_expired'
    detail_msg = (f"{session.get('user_name')} auto-submitted '{qs_row['name'] if qs_row else session_id}' "
                  f"— {'integrity violations' if reason == 'cheat' else 'time expired'}")
    log_action(conn, action_label, category='user',
               entity_type='session', entity_id=session_id,
               entity_name=session.get('user_name'), details=detail_msg)
    conn.commit()
    close_db(conn)
    if reason == 'cheat':
        flash('🚩 Your quiz was automatically submitted due to multiple integrity violations.', 'error')
    else:
        flash('⏰ Time is up! Your answers have been submitted.', 'error')
    return redirect(url_for('results', session_id=session_id))

@app.route('/api/session-status/<int:session_id>')
@login_required
def api_session_status(session_id):
    """Returns whether a session is open for starting right now (used by frontend countdown)."""
    from flask import jsonify
    conn = get_db()
    qs = _fetchone(conn,
        'SELECT id, scheduled_start, is_active FROM quiz_sessions WHERE id=%s', (session_id,)
    )
    close_db(conn)
    if not qs or not qs['is_active']:
        return jsonify({'open': False, 'reason': 'inactive'})
    if qs['scheduled_start']:
        sched = coerce_dt(qs['scheduled_start'])
        seconds_until = int((sched - now_eat()).total_seconds())
        if seconds_until > 0:
            return jsonify({'open': False, 'reason': 'not_yet', 'seconds_until': seconds_until})
    return jsonify({'open': True})


@app.route('/api/cheat/<int:session_id>', methods=['POST'])
@login_required
def cheat_flag(session_id):
    """Record a cheating violation for the current user's active session."""
    violation = request.json.get('violation', 'unknown') if request.is_json else 'unknown'
    # sanitize
    allowed = {'tab_switch', 'window_blur', 'copy_attempt', 'right_click',
               'keyboard_shortcut', 'devtools', 'context_menu', 'auto_submit'}
    violation = violation if violation in allowed else 'unknown'
    conn = get_db()
    us = _fetchone(conn,
        'SELECT id FROM user_sessions WHERE user_id=%s AND session_id=%s AND completed_at IS NULL',
        (session['user_id'], session_id)
    )
    if us:
        _exec(conn,
            'INSERT INTO cheat_flags (user_session_id, violation_type) VALUES (%s,%s)',
            (us['id'], violation)
        )
        # Count total flags for this session
        count = _fetchone(conn,
            'SELECT COUNT(*) as n FROM cheat_flags WHERE user_session_id=%s', (us['id'],)
        )['n']
        qs_row = _fetchone(conn, 'SELECT name FROM quiz_sessions WHERE id=%s', (session_id,))
        log_action(conn, 'cheat_flag', category='user',
                   entity_type='session', entity_id=session_id,
                   entity_name=session.get('user_name'),
                   details=f"{session.get('user_name')} — {violation} in '{qs_row['name'] if qs_row else session_id}' (flag #{count})")
        conn.commit()
        close_db(conn)
        return {'ok': True, 'total_flags': count}
    close_db(conn)
    return {'ok': False}, 404

@app.route('/results', defaults={'session_id': None})
@app.route('/results/<int:session_id>')
@login_required
def results(session_id):
    conn = get_db()
    real_user = _fetchone(conn, 'SELECT id FROM users WHERE id=%s', (session['user_id'],))
    if not real_user:
        close_db(conn)
        session.clear()
        flash('Your session has expired. Please log in again.', 'error')
        return redirect(url_for('index'))
    if session_id:
        us = _fetchone(conn, '''
            SELECT us.*, qs.name as session_name
            FROM user_sessions us JOIN quiz_sessions qs ON us.session_id=qs.id
            WHERE us.user_id=%s AND us.session_id=%s
            ORDER BY us.started_at DESC LIMIT 1
        ''', (session['user_id'], session_id))
        if not us:
            close_db(conn)
            return redirect(url_for('quiz_home'))
        answers = _fetchall(conn, '''
            SELECT ua.*, q.question_text, q.correct_answer, q.option_a, q.option_b, q.option_c, q.option_d,
                   q.points, q.question_type, s.name as section_name
            FROM user_answers ua
            JOIN questions q ON ua.question_id=q.id
            JOIN sections s ON q.section_id=s.id
            WHERE ua.user_session_id=%s
            ORDER BY ua.answered_at
        ''', (us['id'],))
        correct = sum(1 for a in answers if a['is_correct'])
        pts     = sum(float(a['points_earned'] or 0) for a in answers)
        close_db(conn)
        return render_template('results.html', single=True, user_sess=us,
                               answers=answers, correct_count=correct, total_points=pts)
    else:
        all_sessions = _fetchall(conn, '''
            SELECT us.id, us.started_at, us.completed_at, qs.name as session_name,
                   COUNT(ua.id) as total_answered,
                   SUM(CASE WHEN ua.is_correct = 1 THEN 1 ELSE 0 END) as correct_count,
                   SUM(COALESCE(ua.points_earned, 0)) as total_points,
                   us.session_id
            FROM user_sessions us
            JOIN quiz_sessions qs ON us.session_id=qs.id
            LEFT JOIN user_answers ua ON us.id=ua.user_session_id
            LEFT JOIN questions q ON ua.question_id=q.id
            WHERE us.user_id=%s
            GROUP BY us.id, us.started_at, us.completed_at, qs.name, us.session_id
            ORDER BY us.started_at DESC
        ''', (session['user_id'],))
        close_db(conn)
        return render_template('results.html', single=False, all_sessions=all_sessions)

@app.route('/logout')
def logout():
    if 'user_id' in session:
        conn = get_db()
        log_action(conn, 'user_logout', category='user',
                   entity_type='user', entity_id=session['user_id'],
                   entity_name=session.get('user_name'),
                   details=f"{session.get('user_name')} logged out")
        conn.commit()
        close_db(conn)
    session.clear()
    return redirect(url_for('index'))

# ── Timer API ──────────────────────────────────────────────────────────────────
@app.route('/api/timer/<int:session_id>')
@login_required
def api_timer(session_id):
    """Returns the authoritative remaining seconds from the backend."""
    from flask import jsonify
    conn = get_db()
    qs = _fetchone(conn, 'SELECT time_limit_minutes FROM quiz_sessions WHERE id=%s', (session_id,))
    us = _fetchone(conn,
        'SELECT * FROM user_sessions WHERE user_id=%s AND session_id=%s AND completed_at IS NULL',
        (session['user_id'], session_id)
    )
    close_db(conn)
    if not qs or not us:
        return jsonify({'remaining': 0, 'expired': True})
    time_limit = qs['time_limit_minutes'] or 0
    if not time_limit:
        return jsonify({'remaining': None, 'expired': False})
    remaining = get_remaining_seconds(us, time_limit)
    return jsonify({'remaining': remaining, 'expired': remaining <= 0})

# ═══════════════════════════════════════════════════════════════════════════════
#  ADMIN ROUTES
# ═══════════════════════════════════════════════════════════════════════════════

@app.route('/admin', methods=['GET', 'POST'])
def admin_login():
    if session.get('is_admin'):
        return redirect(url_for('admin_dashboard'))
    if request.method == 'POST':
        pw = request.form.get('password', '')
        conn = get_db()
        stored = _fetchone(conn, "SELECT value FROM app_settings WHERE key='admin_password'")
        if stored and pw == stored['value']:
            session['is_admin'] = True
            log_action(conn, 'admin_login', details='Admin logged in')
            conn.commit()
            close_db(conn)
            return redirect(url_for('admin_dashboard'))
        log_action(conn, 'admin_login_failed', details='Failed login attempt')
        conn.commit()
        close_db(conn)
        flash('Wrong password.', 'error')
    return render_template('admin/login.html')

@app.route('/admin/logout')
def admin_logout():
    conn = get_db()
    log_action(conn, 'admin_logout', details='Admin logged out')
    conn.commit()
    close_db(conn)
    session.pop('is_admin', None)
    return redirect(url_for('admin_login'))

@app.route('/admin/dashboard')
@admin_required
def admin_dashboard():
    conn = get_db()
    stats = dict(
        users     = _fetchone(conn, 'SELECT COUNT(*) FROM users')['count'],
        sessions  = _fetchone(conn, 'SELECT COUNT(*) FROM quiz_sessions')['count'],
        questions = _fetchone(conn, 'SELECT COUNT(*) FROM questions')['count'],
        correct   = _fetchone(conn, 'SELECT COUNT(*) FROM user_answers WHERE is_correct=1')['count'],
    )
    leaderboard = _fetchall(conn, '''
        SELECT u.name, u.phone,
               COUNT(DISTINCT us.session_id) as sessions_taken,
               SUM(COALESCE(ua.points_earned, 0)) as total_points,
               SUM(CASE WHEN ua.is_correct = 1 THEN 1 ELSE 0 END) as correct_count
        FROM users u
        LEFT JOIN user_sessions us ON u.id=us.user_id
        LEFT JOIN user_answers ua ON us.id=ua.user_session_id
        LEFT JOIN questions q ON ua.question_id=q.id
        GROUP BY u.id, u.name, u.phone ORDER BY total_points DESC LIMIT 15
    ''')
    close_db(conn)
    return render_template('admin/dashboard.html', stats=stats, leaderboard=leaderboard)

# Sessions CRUD
@app.route('/admin/sessions', methods=['GET', 'POST'])
@admin_required
def admin_sessions():
    conn = get_db()
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'create':
            sched_val = parse_scheduled_start(request.form.get('scheduled_start', ''))
            _exec(conn, 'INSERT INTO quiz_sessions (name, description, randomize_questions, time_limit_minutes, scheduled_start) VALUES (%s,%s,%s,%s,%s)',
                         (request.form['name'], request.form.get('description',''),
                          1 if request.form.get('randomize') else 0,
                          int(request.form.get('time_limit_minutes') or 0),
                          sched_val))
            log_action(conn, 'create_session', entity_type='session',
                       entity_name=request.form['name'],
                       details=f"Created session '{request.form['name']}'")
            conn.commit(); flash('Session created!', 'success')
        elif action == 'toggle_active':
            sid = request.form['sid']
            cur = _exec(conn, 'UPDATE quiz_sessions SET is_active = CASE WHEN is_active=1 THEN 0 ELSE 1 END WHERE id=%s RETURNING name, is_active', (sid,))
            row = cur.fetchone()
            new_state = 'activated' if row and row['is_active'] else 'deactivated'
            log_action(conn, 'toggle_session_active', entity_type='session',
                       entity_id=int(sid), entity_name=row['name'] if row else None,
                       details=f"Session {new_state}")
            conn.commit()
        elif action == 'toggle_randomize':
            sid = request.form['sid']
            cur = _exec(conn, 'UPDATE quiz_sessions SET randomize_questions = CASE WHEN randomize_questions=1 THEN 0 ELSE 1 END WHERE id=%s RETURNING name, randomize_questions', (sid,))
            row = cur.fetchone()
            new_state = 'on' if row and row['randomize_questions'] else 'off'
            log_action(conn, 'toggle_session_randomize', entity_type='session',
                       entity_id=int(sid), entity_name=row['name'] if row else None,
                       details=f"Randomize turned {new_state}")
            conn.commit()
        elif action == 'delete':
            sid = request.form['sid']
            row = _fetchone(conn, 'SELECT name FROM quiz_sessions WHERE id=%s', (sid,))
            _exec(conn, 'DELETE FROM quiz_sessions WHERE id=%s', (sid,))
            log_action(conn, 'delete_session', entity_type='session',
                       entity_id=int(sid), entity_name=row['name'] if row else None,
                       details=f"Deleted session '{row['name'] if row else sid}'")
            conn.commit(); flash('Session deleted.', 'success')
        elif action == 'edit':
            sid = request.form['sid']
            sched_val = parse_scheduled_start(request.form.get('scheduled_start', ''))
            _exec(conn, 'UPDATE quiz_sessions SET name=%s, description=%s, time_limit_minutes=%s, scheduled_start=%s WHERE id=%s',
                         (request.form['name'], request.form.get('description',''),
                          int(request.form.get('time_limit_minutes') or 0),
                          sched_val, sid))
            log_action(conn, 'edit_session', entity_type='session',
                       entity_id=int(sid), entity_name=request.form['name'],
                       details=f"Edited session '{request.form['name']}'")
            conn.commit(); flash('Session updated!', 'success')

    sessions_list = _fetchall(conn, '''
        SELECT qs.id, qs.name, qs.description, qs.is_active, qs.randomize_questions,
               qs.time_limit_minutes, qs.scheduled_start, qs.created_at,
               COUNT(DISTINCT s.id)       as section_count,
               COUNT(DISTINCT q.id)       as question_count,
               COUNT(DISTINCT us.user_id) as participant_count
        FROM quiz_sessions qs
        LEFT JOIN sections s ON qs.id=s.session_id
        LEFT JOIN questions q ON s.id=q.section_id
        LEFT JOIN user_sessions us ON qs.id=us.session_id
        GROUP BY qs.id, qs.name, qs.description, qs.is_active, qs.randomize_questions,
                 qs.time_limit_minutes, qs.scheduled_start, qs.created_at
        ORDER BY qs.created_at DESC
    ''')
    close_db(conn)
    return render_template('admin/sessions.html', sessions=sessions_list)

# Sections CRUD
@app.route('/admin/sessions/<int:session_id>/sections', methods=['GET', 'POST'])
@admin_required
def admin_sections(session_id):
    conn = get_db()
    qs = _fetchone(conn, 'SELECT * FROM quiz_sessions WHERE id=%s', (session_id,))
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'create':
            _exec(conn, 'INSERT INTO sections (session_id, name, order_num) VALUES (%s,%s,%s)',
                         (session_id, request.form['name'], request.form.get('order_num', 0)))
            log_action(conn, 'create_section', entity_type='section',
                       entity_name=request.form['name'],
                       details=f"Created section '{request.form['name']}' in session #{session_id} ({qs['name'] if qs else ''})")
            conn.commit(); flash('Section created!', 'success')
        elif action == 'delete':
            sec_id = request.form['sec_id']
            row = _fetchone(conn, 'SELECT name FROM sections WHERE id=%s', (sec_id,))
            _exec(conn, 'DELETE FROM sections WHERE id=%s', (sec_id,))
            log_action(conn, 'delete_section', entity_type='section',
                       entity_id=int(sec_id), entity_name=row['name'] if row else None,
                       details=f"Deleted section from session '{qs['name'] if qs else session_id}'")
            conn.commit()
        elif action == 'edit':
            sec_id = request.form['sec_id']
            _exec(conn, 'UPDATE sections SET name=%s, order_num=%s WHERE id=%s',
                         (request.form['name'], request.form.get('order_num',0), sec_id))
            log_action(conn, 'edit_section', entity_type='section',
                       entity_id=int(sec_id), entity_name=request.form['name'],
                       details=f"Edited section in session '{qs['name'] if qs else session_id}'")
            conn.commit(); flash('Section updated!', 'success')

    sections_list = _fetchall(conn, '''
        SELECT s.id, s.session_id, s.name, s.order_num, COUNT(q.id) as question_count
        FROM sections s LEFT JOIN questions q ON s.id=q.section_id
        WHERE s.session_id=%s
        GROUP BY s.id, s.session_id, s.name, s.order_num
        ORDER BY s.order_num
    ''', (session_id,))
    close_db(conn)
    return render_template('admin/sections.html', quiz_session=qs, sections=sections_list)

# Questions CRUD
@app.route('/admin/sections/<int:section_id>/questions', methods=['GET', 'POST'])
@admin_required
def admin_questions(section_id):
    conn = get_db()
    sec = _fetchone(conn, '''
        SELECT s.*, qs.name as session_name, qs.id as session_id
        FROM sections s JOIN quiz_sessions qs ON s.session_id=qs.id WHERE s.id=%s
    ''', (section_id,))
    if request.method == 'POST':
        action = request.form.get('action')
        qtype  = request.form.get('question_type', 'single')

        def extract_blank_options():
            """Pull blank_N_options fields and return JSON."""
            bo = []
            i = 0
            while True:
                key = f'blank_{i}_options'
                val = request.form.get(key, '').strip()
                if not val and i > 0:
                    break
                if val:
                    # split by newline or comma
                    opts = [o.strip() for o in val.replace('\n', ',').split(',') if o.strip()]
                    bo.append(opts)
                    i += 1
                else:
                    break
            return json.dumps(bo)

        def extract_correct():
            if qtype == 'multi':
                checked = request.form.getlist('correct_answer')
                return ','.join(sorted(x.upper() for x in checked))
            elif qtype == 'fill_blank':
                bo = json.loads(extract_blank_options() or '[]')
                parts = [request.form.get(f'blank_{i}_correct', '').strip() for i in range(len(bo))]
                return '|'.join(parts)
            else:
                return request.form.get('correct_answer', '').upper()

        if action == 'create':
            bo_json = extract_blank_options() if qtype == 'fill_blank' else '[]'
            correct = request.form.get('correct_answer_hidden') or extract_correct()
            if qtype == 'fill_blank':
                # Re-extract properly
                bo = []
                i = 0
                while True:
                    opts_raw = request.form.get(f'blank_{i}_options', '').strip()
                    if not opts_raw:
                        break
                    opts = [o.strip() for o in opts_raw.replace('\n',',').split(',') if o.strip()]
                    bo.append(opts)
                    i += 1
                bo_json = json.dumps(bo)
                parts = [request.form.get(f'blank_{i}_correct', '').strip() for i in range(len(bo))]
                correct = '|'.join(parts)
            elif qtype == 'multi':
                checked = request.form.getlist('correct_answer')
                correct = ','.join(sorted(x.upper() for x in checked))
            else:
                correct = request.form.get('correct_answer', '').upper()
                bo_json = '[]'

            _exec(conn, '''
                INSERT INTO questions (section_id, question_type, question_text,
                    option_a, option_b, option_c, option_d,
                    correct_answer, blank_options, points, order_num)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            ''', (section_id, qtype, request.form['question_text'],
                  request.form.get('option_a',''), request.form.get('option_b',''),
                  request.form.get('option_c',''), request.form.get('option_d',''),
                  correct, bo_json,
                  request.form.get('points', 1), request.form.get('order_num', 0)))
            log_action(conn, 'create_question', entity_type='question',
                       entity_name=request.form['question_text'][:80],
                       details=f"Added {qtype} question to section '{sec['name'] if sec else section_id}'")
            conn.commit(); flash('Question added!', 'success')

        elif action == 'delete':
            q_id_del = request.form['q_id']
            qrow = _fetchone(conn, 'SELECT question_text FROM questions WHERE id=%s', (q_id_del,))
            _exec(conn, 'DELETE FROM questions WHERE id=%s', (q_id_del,))
            log_action(conn, 'delete_question', entity_type='question',
                       entity_id=int(q_id_del),
                       entity_name=qrow['question_text'][:80] if qrow else None,
                       details=f"Deleted from section '{sec['name'] if sec else section_id}'")
            conn.commit()

        elif action == 'edit':
            q_id_edit = request.form['q_id']
            qtype_edit = request.form.get('question_type', 'single')
            if qtype_edit == 'fill_blank':
                bo = []
                i = 0
                while True:
                    opts_raw = request.form.get(f'blank_{i}_options', '').strip()
                    if not opts_raw:
                        break
                    opts = [o.strip() for o in opts_raw.replace('\n',',').split(',') if o.strip()]
                    bo.append(opts)
                    i += 1
                bo_json = json.dumps(bo)
                parts = [request.form.get(f'blank_{i}_correct', '').strip() for i in range(len(bo))]
                correct_edit = '|'.join(parts)
            elif qtype_edit == 'multi':
                checked = request.form.getlist('correct_answer')
                correct_edit = ','.join(sorted(x.upper() for x in checked))
                bo_json = '[]'
            else:
                correct_edit = request.form.get('correct_answer','').upper()
                bo_json = '[]'

            _exec(conn, '''
                UPDATE questions SET question_type=%s, question_text=%s,
                    option_a=%s, option_b=%s, option_c=%s, option_d=%s,
                    correct_answer=%s, blank_options=%s, points=%s, order_num=%s
                WHERE id=%s
            ''', (qtype_edit, request.form['question_text'],
                  request.form.get('option_a',''), request.form.get('option_b',''),
                  request.form.get('option_c',''), request.form.get('option_d',''),
                  correct_edit, bo_json,
                  request.form.get('points',1), request.form.get('order_num',0),
                  q_id_edit))
            log_action(conn, 'edit_question', entity_type='question',
                       entity_id=int(q_id_edit),
                       entity_name=request.form['question_text'][:80],
                       details=f"Edited {qtype_edit} question in section '{sec['name'] if sec else section_id}'")
            conn.commit(); flash('Question updated!', 'success')

    questions_list = [dict(q) for q in _fetchall(conn,
        'SELECT * FROM questions WHERE section_id=%s ORDER BY order_num', (section_id,)
    )]
    close_db(conn)
    return render_template('admin/questions.html', section=sec, questions=questions_list)

# Users & scores
@app.route('/admin/users')
@admin_required
def admin_users():
    conn = get_db()
    users = _fetchall(conn, '''
        SELECT u.id, u.name, u.phone, u.created_at,
               COUNT(DISTINCT us.session_id) as sessions_taken,
               SUM(COALESCE(ua.points_earned, 0)) as total_points,
               SUM(CASE WHEN ua.is_correct = 1 THEN 1 ELSE 0 END) as correct_count,
               COUNT(ua.id) as total_answered,
               (SELECT COUNT(*) FROM cheat_flags cf
                JOIN user_sessions us2 ON cf.user_session_id=us2.id
                WHERE us2.user_id=u.id) as cheat_count
        FROM users u
        LEFT JOIN user_sessions us ON u.id=us.user_id
        LEFT JOIN user_answers ua ON us.id=ua.user_session_id
        LEFT JOIN questions q ON ua.question_id=q.id
        GROUP BY u.id, u.name, u.phone, u.created_at
        ORDER BY total_points DESC NULLS LAST
    ''')
    close_db(conn)
    return render_template('admin/users.html', users=users)

@app.route('/admin/users/<int:user_id>')
@admin_required
def admin_user_detail(user_id):
    conn = get_db()
    user = _fetchone(conn, 'SELECT * FROM users WHERE id=%s', (user_id,))
    sessions_data = _fetchall(conn, '''
        SELECT us.id, us.user_id, us.session_id, us.started_at, us.completed_at,
               qs.name as session_name,
               COUNT(ua.id) as total_answered,
               SUM(CASE WHEN ua.is_correct = 1 THEN 1 ELSE 0 END) as correct_count,
               SUM(COALESCE(ua.points_earned, 0)) as total_points,
               (SELECT COUNT(*) FROM questions qq JOIN sections ss ON qq.section_id=ss.id WHERE ss.session_id=qs.id) as total_questions
        FROM user_sessions us JOIN quiz_sessions qs ON us.session_id=qs.id
        LEFT JOIN user_answers ua ON us.id=ua.user_session_id
        LEFT JOIN questions q ON ua.question_id=q.id
        WHERE us.user_id=%s
        GROUP BY us.id, us.user_id, us.session_id, us.started_at, us.completed_at, qs.id, qs.name
        ORDER BY us.started_at DESC
    ''', (user_id,))
    codes = _fetchall(conn, '''
        SELECT ua.reward_code, ua.answered_at, q.question_text, qs.name as session_name
        FROM user_answers ua
        JOIN questions q ON ua.question_id=q.id
        JOIN sections s ON q.section_id=s.id
        JOIN quiz_sessions qs ON s.session_id=qs.id
        JOIN user_sessions us ON ua.user_session_id=us.id
        WHERE us.user_id=%s AND ua.is_correct=1
        ORDER BY ua.answered_at DESC
    ''', (user_id,))
    # Cheat flags
    cheat_flags = _fetchall(conn, '''
        SELECT cf.violation_type, cf.flagged_at, qs.name as session_name
        FROM cheat_flags cf
        JOIN user_sessions us ON cf.user_session_id=us.id
        JOIN quiz_sessions qs ON us.session_id=qs.id
        WHERE us.user_id=%s
        ORDER BY cf.flagged_at DESC LIMIT 50
    ''', (user_id,))
    close_db(conn)
    return render_template('admin/user_detail.html', user=user, sessions_data=sessions_data,
                           codes=codes, cheat_flags=cheat_flags)

@app.route('/admin/performance')
@admin_required
def admin_performance():
    conn = get_db()
    session_id = request.args.get('session_id', type=int)

    # All sessions for the filter dropdown
    all_sessions = _fetchall(conn,
        'SELECT id, name FROM quiz_sessions ORDER BY created_at DESC'
    )

    if not session_id and all_sessions:
        session_id = all_sessions[0]['id']

    perf = None
    q_stats = []
    section_stats = []
    top_users = []
    score_dist = {}

    if session_id:
        perf = _fetchone(conn, '''
            SELECT qs.id, qs.name, qs.description, qs.time_limit_minutes,
                   qs.randomize_questions, qs.scheduled_start,
                   COUNT(DISTINCT us.user_id)             as participant_count,
                   COUNT(DISTINCT CASE WHEN us.completed_at IS NOT NULL THEN us.user_id END) as completed_count,
                   COUNT(DISTINCT CASE WHEN us.completed_at IS NULL     THEN us.user_id END) as inprogress_count,
                   COUNT(DISTINCT q.id)                   as total_questions,
                   SUM(CASE WHEN ua.is_correct = 1 THEN 1 ELSE 0 END)  as total_correct,
                   COUNT(ua.id)                           as total_answered,
                   AVG(CASE WHEN us.completed_at IS NOT NULL
                       THEN (SELECT SUM(COALESCE(ua2.points_earned, 0))
                             FROM user_answers ua2 JOIN questions q2 ON ua2.question_id=q2.id
                             WHERE ua2.user_session_id=us.id) END) as avg_score
            FROM quiz_sessions qs
            LEFT JOIN user_sessions us ON qs.id=us.session_id
            LEFT JOIN user_answers ua  ON us.id=ua.user_session_id
            LEFT JOIN sections s       ON s.session_id=qs.id
            LEFT JOIN questions q      ON q.section_id=s.id AND q.id IS NOT NULL
            WHERE qs.id=%s
            GROUP BY qs.id, qs.name, qs.description, qs.time_limit_minutes,
                     qs.randomize_questions, qs.scheduled_start
        ''', (session_id,))

        # Per-question stats
        q_stats = _fetchall(conn, '''
            SELECT q.id, q.question_text, q.question_type, q.points,
                   sec.name as section_name,
                   COUNT(ua.id)                                     as attempts,
                   SUM(CASE WHEN ua.is_correct = 1 THEN 1 ELSE 0 END)  as correct,
                   ROUND(100.0 * SUM(CASE WHEN ua.is_correct = 1 THEN 1 ELSE 0 END)
                         / NULLIF(COUNT(ua.id), 0), 1)              as pct_correct
            FROM questions q
            JOIN sections sec ON q.section_id=sec.id
            LEFT JOIN user_answers ua ON ua.question_id=q.id
                AND ua.user_session_id IN (
                    SELECT id FROM user_sessions WHERE session_id=%s
                )
            WHERE sec.session_id=%s
            GROUP BY q.id, q.question_text, q.question_type, q.points, sec.name
            ORDER BY pct_correct ASC, attempts DESC
        ''', (session_id, session_id))

        # Per-section stats
        section_stats = _fetchall(conn, '''
            SELECT sec.name,
                   COUNT(DISTINCT q.id)                                         as q_count,
                   SUM(CASE WHEN ua.is_correct = 1 THEN 1 ELSE 0 END)              as correct,
                   COUNT(ua.id)                                                 as answered,
                   ROUND(100.0 * SUM(CASE WHEN ua.is_correct = 1 THEN 1 ELSE 0 END)
                         / NULLIF(COUNT(ua.id), 0), 1)                         as pct_correct
            FROM sections sec
            LEFT JOIN questions q  ON q.section_id=sec.id
            LEFT JOIN user_answers ua ON ua.question_id=q.id
                AND ua.user_session_id IN (SELECT id FROM user_sessions WHERE session_id=%s)
            WHERE sec.session_id=%s
            GROUP BY sec.id, sec.name, sec.order_num
            ORDER BY sec.order_num
        ''', (session_id, session_id))

        # All participants for this session (for reset table + integrity flags)
        top_users = _fetchall(conn, '''
            SELECT u.id as user_id, u.name, u.phone,
                   SUM(COALESCE(ua.points_earned, 0)) as points,
                   SUM(CASE WHEN ua.is_correct = 1 THEN 1 ELSE 0 END)        as correct,
                   COUNT(ua.id)                                               as answered,
                   us.completed_at,
                   us.started_at,
                   (SELECT COUNT(*) FROM cheat_flags cf
                    WHERE cf.user_session_id = us.id)                         as cheat_count
            FROM user_sessions us
            JOIN users u ON us.user_id=u.id
            LEFT JOIN user_answers ua ON ua.user_session_id=us.id
            LEFT JOIN questions q     ON ua.question_id=q.id
            WHERE us.session_id=%s
            GROUP BY us.id, u.id, u.name, u.phone, us.completed_at, us.started_at
            ORDER BY points DESC, correct DESC
        ''', (session_id,))

        # Score distribution buckets: 0-20, 21-40, 41-60, 61-80, 81-100 %
        all_scores = _fetchall(conn, '''
            SELECT ROUND(100.0 * SUM(CASE WHEN ua.is_correct = 1 THEN 1 ELSE 0 END)
                         / NULLIF(COUNT(ua.id),0)) as pct
            FROM user_sessions us
            LEFT JOIN user_answers ua ON ua.user_session_id=us.id
            WHERE us.session_id=%s AND us.completed_at IS NOT NULL
            GROUP BY us.id
        ''', (session_id,))
        buckets = {'0–20': 0, '21–40': 0, '41–60': 0, '61–80': 0, '81–100': 0}
        for row in all_scores:
            p = row['pct'] or 0
            if   p <= 20:  buckets['0–20']   += 1
            elif p <= 40:  buckets['21–40']  += 1
            elif p <= 60:  buckets['41–60']  += 1
            elif p <= 80:  buckets['61–80']  += 1
            else:          buckets['81–100'] += 1
        score_dist = buckets

    close_db(conn)
    return render_template('admin/performance.html',
                           all_sessions=all_sessions,
                           selected_id=session_id,
                           perf=perf,
                           q_stats=q_stats,
                           section_stats=section_stats,
                           top_users=top_users,
                           score_dist=score_dist)


@app.route('/admin/performance/export')
@admin_required
def export_performance():
    """Export per-user or per-question results for a session as Excel (.xlsx)."""
    import io
    from flask import Response
    import openpyxl
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.utils import get_column_letter

    session_id  = request.args.get('session_id', type=int)
    export_type = request.args.get('type', 'users')   # 'users' or 'questions'

    if not session_id:
        flash('No session selected.', 'error')
        return redirect(url_for('admin_performance'))

    conn    = get_db()
    qs_row  = _fetchone(conn, 'SELECT name FROM quiz_sessions WHERE id=%s', (session_id,))
    if not qs_row:
        close_db(conn)
        flash('Session not found.', 'error')
        return redirect(url_for('admin_performance'))

    safe_name = qs_row['name'].replace(' ', '_')

    # ── Shared style helpers ──────────────────────────────────────────────────
    def _hdr(text, bg='1E3A5F', fg='FFFFFF', bold=True, size=11):
        c = openpyxl.cell.cell.WriteOnlyCell if False else None
        pass

    HDR_FILL   = PatternFill('solid', fgColor='1E3A5F')
    HDR_FONT   = Font(bold=True, color='FFFFFF', size=11)
    ALT_FILL   = PatternFill('solid', fgColor='F0F4FA')
    GREEN_FILL = PatternFill('solid', fgColor='D1FAE5')
    RED_FILL   = PatternFill('solid', fgColor='FEE2E2')
    CENTER     = Alignment(horizontal='center', vertical='center', wrap_text=True)
    LEFT       = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    thin       = Side(style='thin', color='D1D5DB')
    BORDER     = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header_row(ws, row_num, col_count):
        for c in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=c)
            cell.fill   = HDR_FILL
            cell.font   = HDR_FONT
            cell.border = BORDER
            cell.alignment = CENTER

    def style_data_row(ws, row_num, col_count, alt=False, fill=None):
        for c in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=c)
            cell.fill   = fill if fill else (ALT_FILL if alt else PatternFill())
            cell.border = BORDER
            cell.alignment = LEFT

    wb = openpyxl.Workbook()

    # ── PARTICIPANTS sheet ────────────────────────────────────────────────────
    if export_type == 'users':
        ws = wb.active
        ws.title = 'Participants'
        ws.freeze_panes = 'A2'
        ws.row_dimensions[1].height = 28

        headers = ['#', 'Name', 'Phone', 'Total Points', 'Correct',
                   'Wrong', 'Answered', 'Accuracy %', 'Integrity Flags',
                   'Started', 'Completed', 'Status']
        ws.append(headers)
        style_header_row(ws, 1, len(headers))

        rows = _fetchall(conn, '''
            SELECT u.name, u.phone,
                   SUM(COALESCE(ua.points_earned, 0)) as total_points,
                   SUM(CASE WHEN ua.is_correct = 1 THEN 1 ELSE 0 END)        as correct,
                   COUNT(ua.id) - SUM(CASE WHEN ua.is_correct = 1 THEN 1 ELSE 0 END) as wrong,
                   COUNT(ua.id)                                               as answered,
                   ROUND(100.0 * SUM(CASE WHEN ua.is_correct = 1 THEN 1 ELSE 0 END)
                         / NULLIF(COUNT(ua.id),0), 1)                        as accuracy,
                   us.started_at, us.completed_at,
                   (SELECT COUNT(*) FROM cheat_flags cf
                    WHERE cf.user_session_id = us.id)                        as integrity_flags
            FROM user_sessions us
            JOIN users u ON us.user_id=u.id
            LEFT JOIN user_answers ua ON ua.user_session_id=us.id
            LEFT JOIN questions q     ON ua.question_id=q.id
            WHERE us.session_id=%s
            GROUP BY us.id, u.name, u.phone, us.started_at, us.completed_at
            ORDER BY total_points DESC
        ''', (session_id,))

        for i, r in enumerate(rows, start=1):
            status = 'Completed' if r['completed_at'] else 'In Progress'
            row_data = [
                i, r['name'], r['phone'],
                int(r['total_points'] or 0),
                int(r['correct'] or 0),
                int(r['wrong'] or 0),
                int(r['answered'] or 0),
                float(r['accuracy'] or 0),
                int(r['integrity_flags'] or 0),
                str(r['started_at'])[:16] if r['started_at'] else '',
                str(r['completed_at'])[:16] if r['completed_at'] else '',
                status,
            ]
            ws.append(row_data)
            fill = GREEN_FILL if status == 'Completed' else None
            style_data_row(ws, i + 1, len(headers), alt=(i % 2 == 0), fill=fill)

        # Column widths
        for col, width in zip(range(1, len(headers)+1),
                              [5, 22, 15, 13, 10, 10, 11, 12, 16, 18, 18, 12]):
            ws.column_dimensions[get_column_letter(col)].width = width

        filename = f'{safe_name}_participants.xlsx'

    # ── QUESTIONS sheet ───────────────────────────────────────────────────────
    else:
        ws = wb.active
        ws.title = 'Question Stats'
        ws.freeze_panes = 'A2'
        ws.row_dimensions[1].height = 28

        headers = ['#', 'Section', 'Question', 'Type', 'Points',
                   'Attempts', 'Correct', 'Wrong', '% Correct']
        ws.append(headers)
        style_header_row(ws, 1, len(headers))

        rows = _fetchall(conn, '''
            SELECT sec.name as section, q.question_text, q.question_type, q.points,
                   COUNT(ua.id) as attempts,
                   SUM(CASE WHEN ua.is_correct = 1 THEN 1 ELSE 0 END) as correct,
                   COUNT(ua.id) - SUM(CASE WHEN ua.is_correct = 1 THEN 1 ELSE 0 END) as wrong,
                   ROUND(100.0 * SUM(CASE WHEN ua.is_correct = 1 THEN 1 ELSE 0 END)
                         / NULLIF(COUNT(ua.id),0), 1) as pct_correct
            FROM questions q
            JOIN sections sec ON q.section_id=sec.id
            LEFT JOIN user_answers ua ON ua.question_id=q.id
                AND ua.user_session_id IN (SELECT id FROM user_sessions WHERE session_id=%s)
            WHERE sec.session_id=%s
            GROUP BY q.id, q.question_text, q.question_type, q.points,
                     sec.name, sec.order_num
            ORDER BY sec.order_num, q.order_num
        ''', (session_id, session_id))

        for i, r in enumerate(rows, start=1):
            pct = float(r['pct_correct'] or 0)
            ws.append([i, r['section'], r['question_text'], r['question_type'],
                       int(r['points'] or 0), int(r['attempts'] or 0),
                       int(r['correct'] or 0), int(r['wrong'] or 0), pct])
            fill = GREEN_FILL if pct >= 70 else (RED_FILL if pct < 40 else None)
            style_data_row(ws, i + 1, len(headers), alt=(i % 2 == 0), fill=fill)

        for col, width in zip(range(1, len(headers)+1),
                              [5, 18, 50, 10, 8, 10, 10, 8, 12]):
            ws.column_dimensions[get_column_letter(col)].width = width

        filename = f'{safe_name}_questions.xlsx'

    close_db(conn)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )


@app.route('/admin/sessions/<int:session_id>/export')
@admin_required
def export_session_full(session_id):
    """
    Full session export as a multi-sheet Excel workbook:
      Sheet 1 – Session Overview  (name, settings, aggregate stats)
      Sheet 2 – Questions         (all sections & questions with options + correct answer)
      Sheet 3 – Participant Results (each user's answers per question)
      Sheet 4 – Leaderboard       (ranked participants with scores)
    """
    import io
    from flask import Response
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    conn   = get_db()
    qs_row = _fetchone(conn, 'SELECT * FROM quiz_sessions WHERE id=%s', (session_id,))
    if not qs_row:
        close_db(conn)
        flash('Session not found.', 'error')
        return redirect(url_for('admin_sessions'))

    # ── Shared styles ─────────────────────────────────────────────────────────
    HDR_DARK   = PatternFill('solid', fgColor='1E3A5F')   # navy
    HDR_MID    = PatternFill('solid', fgColor='2D6A4F')   # forest green (sub-headers)
    HDR_AMBER  = PatternFill('solid', fgColor='92400E')   # amber-brown
    ALT_FILL   = PatternFill('solid', fgColor='F0F4FA')
    GREEN_FILL = PatternFill('solid', fgColor='D1FAE5')
    RED_FILL   = PatternFill('solid', fgColor='FEE2E2')
    AMBER_FILL = PatternFill('solid', fgColor='FEF3C7')
    WHITE_FILL = PatternFill('solid', fgColor='FFFFFF')
    TITLE_FONT = Font(bold=True, color='FFFFFF', size=12)
    HDR_FONT   = Font(bold=True, color='FFFFFF', size=11)
    BOLD       = Font(bold=True, size=11)
    NORMAL     = Font(size=10)
    CENTER     = Alignment(horizontal='center', vertical='center', wrap_text=True)
    LEFT       = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    RIGHT      = Alignment(horizontal='right',  vertical='center')
    thin       = Side(style='thin',   color='D1D5DB')
    med        = Side(style='medium', color='9CA3AF')
    THIN_BDR   = Border(left=thin, right=thin, top=thin, bottom=thin)
    MED_BDR    = Border(left=med,  right=med,  top=med,  bottom=med)

    def hrow(ws, row_num, col_count, fill=None):
        """Style an entire row as a header."""
        for c in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=c)
            cell.fill      = fill or HDR_DARK
            cell.font      = HDR_FONT
            cell.border    = THIN_BDR
            cell.alignment = CENTER

    def drow(ws, row_num, col_count, alt=False, fill=None):
        for c in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=c)
            cell.fill      = fill if fill else (ALT_FILL if alt else WHITE_FILL)
            cell.font      = NORMAL
            cell.border    = THIN_BDR
            cell.alignment = LEFT

    def col_widths(ws, widths):
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    wb = openpyxl.Workbook()

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 1 – Session Overview
    # ══════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = 'Session Overview'
    ws1.column_dimensions['A'].width = 26
    ws1.column_dimensions['B'].width = 40

    # Title banner
    ws1.merge_cells('A1:B1')
    ws1['A1'] = f'Session Overview – {qs_row["name"]}'
    ws1['A1'].fill = HDR_DARK
    ws1['A1'].font = Font(bold=True, color='FFFFFF', size=14)
    ws1['A1'].alignment = CENTER
    ws1.row_dimensions[1].height = 32

    def info_row(ws, key, val, row, alt=False):
        kc = ws.cell(row=row, column=1, value=key)
        vc = ws.cell(row=row, column=2, value=val)
        bg = ALT_FILL if alt else WHITE_FILL
        kc.fill  = vc.fill  = bg
        kc.font  = Font(bold=True, size=10)
        vc.font  = Font(size=10)
        kc.border = vc.border = THIN_BDR
        kc.alignment = LEFT
        vc.alignment = LEFT

    sess_details = [
        ('Session Name',    qs_row['name']),
        ('Description',     qs_row['description'] or '—'),
        ('Status',          'Active' if qs_row['is_active'] else 'Inactive'),
        ('Randomize Qs',    'Yes' if qs_row['randomize_questions'] else 'No'),
        ('Time Limit',      f"{qs_row['time_limit_minutes']} min" if qs_row['time_limit_minutes'] else 'No Limit'),
        ('Scheduled Start', str(qs_row['scheduled_start'])[:16] if qs_row['scheduled_start'] else 'Immediate'),
        ('Created At',      str(qs_row['created_at'])[:16] if qs_row['created_at'] else '—'),
    ]
    for i, (k, v) in enumerate(sess_details, start=2):
        info_row(ws1, k, v, i, alt=(i % 2 == 0))

    # Stats sub-header
    stats_row = len(sess_details) + 3
    ws1.merge_cells(f'A{stats_row}:B{stats_row}')
    ws1.cell(row=stats_row, column=1, value='📊 Aggregate Statistics').fill = HDR_MID
    ws1.cell(row=stats_row, column=1).font = Font(bold=True, color='FFFFFF', size=11)
    ws1.cell(row=stats_row, column=1).alignment = CENTER
    ws1.row_dimensions[stats_row].height = 22

    agg = _fetchone(conn, '''
        SELECT
            COUNT(DISTINCT us.user_id)                                     as participants,
            COUNT(DISTINCT CASE WHEN us.completed_at IS NOT NULL
                           THEN us.user_id END)                            as completed,
            COUNT(ua.id)                                                   as total_answered,
            SUM(CASE WHEN ua.is_correct = 1 THEN 1 ELSE 0 END)            as total_correct,
            ROUND(100.0 * SUM(CASE WHEN ua.is_correct = 1 THEN 1 ELSE 0 END)
                  / NULLIF(COUNT(ua.id), 0), 1)                            as avg_accuracy,
            (SELECT COUNT(*) FROM sections WHERE session_id=%s)            as section_count,
            (SELECT COUNT(*) FROM questions q2
             JOIN sections s2 ON q2.section_id=s2.id
             WHERE s2.session_id=%s)                                       as question_count
        FROM user_sessions us
        LEFT JOIN user_answers ua ON ua.user_session_id=us.id
        WHERE us.session_id=%s
    ''', (session_id, session_id, session_id))

    stat_details = [
        ('Total Participants',  agg['participants'] or 0),
        ('Completed',           agg['completed'] or 0),
        ('Total Answers',       agg['total_answered'] or 0),
        ('Total Correct',       agg['total_correct'] or 0),
        ('Average Accuracy',    f"{agg['avg_accuracy'] or 0}%"),
        ('Sections',            agg['section_count'] or 0),
        ('Questions',           agg['question_count'] or 0),
    ]
    for i, (k, v) in enumerate(stat_details, start=stats_row + 1):
        info_row(ws1, k, v, i, alt=(i % 2 == 0))

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 2 – Questions (all sections + questions with options + correct answer)
    # ══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet('Questions')
    ws2.freeze_panes = 'A2'
    ws2.row_dimensions[1].height = 26

    q_headers = ['#', 'Section', 'Question', 'Type', 'Points',
                 'Option A', 'Option B', 'Option C', 'Option D',
                 'Correct Answer', 'Attempts', 'Correct', '% Correct']
    ws2.append(q_headers)
    hrow(ws2, 1, len(q_headers))

    questions = _fetchall(conn, '''
        SELECT q.*, sec.name as section_name, sec.order_num as sec_order,
               COUNT(ua.id) as attempts,
               SUM(CASE WHEN ua.is_correct = 1 THEN 1 ELSE 0 END) as correct_count,
               ROUND(100.0 * SUM(CASE WHEN ua.is_correct = 1 THEN 1 ELSE 0 END)
                     / NULLIF(COUNT(ua.id), 0), 1) as pct_correct
        FROM questions q
        JOIN sections sec ON q.section_id = sec.id
        LEFT JOIN user_answers ua ON ua.question_id = q.id
            AND ua.user_session_id IN (SELECT id FROM user_sessions WHERE session_id=%s)
        WHERE sec.session_id = %s
        GROUP BY q.id, sec.name, sec.order_num
        ORDER BY sec.order_num, q.order_num
    ''', (session_id, session_id))

    # Map letter -> full option text for correct answer display
    def expand_correct(q):
        mapping = {
            'A': q['option_a'], 'B': q['option_b'],
            'C': q['option_c'], 'D': q['option_d'],
        }
        parts = []
        for letter in (q['correct_answer'] or '').split(','):
            letter = letter.strip().upper()
            text   = mapping.get(letter, '')
            parts.append(f"{letter}: {text}" if text else letter)
        return ' | '.join(parts)

    for i, q in enumerate(questions, start=1):
        pct  = float(q['pct_correct'] or 0)
        fill = GREEN_FILL if pct >= 70 else (RED_FILL if pct < 40 and q['attempts'] else None)
        ws2.append([
            i,
            q['section_name'],
            q['question_text'],
            q['question_type'],
            int(q['points'] or 0),
            q['option_a'] or '',
            q['option_b'] or '',
            q['option_c'] or '',
            q['option_d'] or '',
            expand_correct(q),
            int(q['attempts'] or 0),
            int(q['correct_count'] or 0),
            pct,
        ])
        drow(ws2, i + 1, len(q_headers), alt=(i % 2 == 0), fill=fill)

    col_widths(ws2, [5, 18, 50, 10, 8, 22, 22, 22, 22, 30, 10, 10, 12])

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 3 – Per-User Detailed Answers
    # ══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet('Participant Results')
    ws3.freeze_panes = 'A2'
    ws3.row_dimensions[1].height = 26

    r_headers = ['#', 'Participant', 'Phone', 'Section', 'Question',
                 'Their Answer', 'Correct Answer', 'Result', 'Points Earned',
                 'Reward Code', 'Answered At']
    ws3.append(r_headers)
    hrow(ws3, 1, len(r_headers), fill=HDR_AMBER)

    detail = _fetchall(conn, '''
        SELECT u.name, u.phone,
               sec.name  as section_name,
               q.question_text, q.question_type,
               q.option_a, q.option_b, q.option_c, q.option_d,
               q.correct_answer,
               ua.selected_answer, ua.is_correct, ua.points_earned, ua.reward_code,
               ua.answered_at,
               q.points
        FROM user_sessions us
        JOIN users u           ON us.user_id = u.id
        JOIN user_answers ua   ON ua.user_session_id = us.id
        JOIN questions q       ON ua.question_id = q.id
        JOIN sections sec      ON q.section_id = sec.id
        WHERE us.session_id = %s
        ORDER BY u.name, sec.order_num, q.order_num, ua.answered_at
    ''', (session_id,))

    def expand_answer(selected, q_row):
        """Convert stored letter(s) to human-readable option text."""
        mapping = {
            'A': q_row['option_a'], 'B': q_row['option_b'],
            'C': q_row['option_c'], 'D': q_row['option_d'],
        }
        qtype = q_row['question_type'] or 'single'
        if qtype == 'fill_blank':
            return selected  # already text
        parts = []
        for letter in selected.split(','):
            letter = letter.strip().upper()
            text   = mapping.get(letter, '')
            parts.append(f"{letter}: {text}" if text else letter)
        return ' | '.join(parts)

    for i, r in enumerate(detail, start=1):
        q_proxy = {
            'option_a': r['option_a'], 'option_b': r['option_b'],
            'option_c': r['option_c'], 'option_d': r['option_d'],
            'question_type': r['question_type'],
        }
        their_ans   = expand_answer(r['selected_answer'], q_proxy)
        correct_ans = expand_answer(r['correct_answer'],  {**q_proxy, 'question_type': 'single'})
        is_correct  = bool(r['is_correct'])
        pts_earned  = float(r['points_earned'] or 0) if r.get('points_earned') is not None else (float(r['points'] or 0) if is_correct else 0.0)

        ws3.append([
            i,
            r['name'],
            r['phone'],
            r['section_name'],
            r['question_text'],
            their_ans,
            correct_ans,
            '✓ Correct' if is_correct else '✗ Wrong',
            pts_earned,
            r['reward_code'] or '',
            str(r['answered_at'])[:16] if r['answered_at'] else '',
        ])
        fill = GREEN_FILL if is_correct else RED_FILL
        drow(ws3, i + 1, len(r_headers), fill=fill)

    col_widths(ws3, [5, 20, 14, 18, 45, 30, 30, 12, 13, 12, 16])

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 4 – Leaderboard
    # ══════════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet('Leaderboard')
    ws4.freeze_panes = 'A2'
    ws4.row_dimensions[1].height = 26

    lb_headers = ['Rank', 'Name', 'Phone', 'Total Points', 'Correct',
                  'Wrong', 'Answered', 'Accuracy %', 'Integrity Flags',
                  'Started', 'Completed', 'Status']
    ws4.append(lb_headers)
    hrow(ws4, 1, len(lb_headers), fill=HDR_MID)

    lb_rows = _fetchall(conn, '''
        SELECT u.name, u.phone,
               SUM(COALESCE(ua.points_earned, 0)) as total_points,
               SUM(CASE WHEN ua.is_correct = 1 THEN 1 ELSE 0 END)        as correct,
               COUNT(ua.id) - SUM(CASE WHEN ua.is_correct = 1 THEN 1 ELSE 0 END) as wrong,
               COUNT(ua.id)                                               as answered,
               ROUND(100.0 * SUM(CASE WHEN ua.is_correct = 1 THEN 1 ELSE 0 END)
                     / NULLIF(COUNT(ua.id),0), 1)                        as accuracy,
               us.started_at, us.completed_at,
               (SELECT COUNT(*) FROM cheat_flags cf
                WHERE cf.user_session_id = us.id)                        as integrity_flags
        FROM user_sessions us
        JOIN users u ON us.user_id = u.id
        LEFT JOIN user_answers ua ON ua.user_session_id = us.id
        LEFT JOIN questions q     ON ua.question_id = q.id
        WHERE us.session_id = %s
        GROUP BY us.id, u.name, u.phone, us.started_at, us.completed_at
        ORDER BY total_points DESC, correct DESC
    ''', (session_id,))

    medals = {1: '🥇', 2: '🥈', 3: '🥉'}
    for i, r in enumerate(lb_rows, start=1):
        status = 'Completed' if r['completed_at'] else 'In Progress'
        rank   = medals.get(i, str(i))
        ws4.append([
            rank,
            r['name'],
            r['phone'],
            int(r['total_points'] or 0),
            int(r['correct'] or 0),
            int(r['wrong'] or 0),
            int(r['answered'] or 0),
            float(r['accuracy'] or 0),
            int(r['integrity_flags'] or 0),
            str(r['started_at'])[:16]  if r['started_at']  else '',
            str(r['completed_at'])[:16] if r['completed_at'] else '',
            status,
        ])
        fill = GREEN_FILL if status == 'Completed' else AMBER_FILL
        drow(ws4, i + 1, len(lb_headers), fill=fill)

    col_widths(ws4, [7, 22, 15, 13, 10, 10, 11, 12, 16, 18, 18, 12])

    close_db(conn)

    # ── Stream the workbook ───────────────────────────────────────────────────
    safe = qs_row['name'].replace(' ', '_')
    buf  = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition':
                 f'attachment; filename="{safe}_full_export.xlsx"'}
    )


@app.route('/admin/performance/participant-answers')
@admin_required
def participant_answers():
    """JSON: returns one participant's full answer breakdown for a session.
    Query params: session_id (int), user_id (int)
    """
    from flask import jsonify
    session_id = request.args.get('session_id', type=int)
    user_id    = request.args.get('user_id',    type=int)
    if not session_id or not user_id:
        return jsonify({'error': 'Missing params'}), 400

    conn = get_db()

    info = _fetchone(conn, '''
        SELECT u.name, u.phone, qs.name as session_name,
               us.started_at, us.completed_at,
               (SELECT COUNT(*) FROM cheat_flags cf WHERE cf.user_session_id = us.id) as flag_count
        FROM user_sessions us
        JOIN users u          ON us.user_id    = u.id
        JOIN quiz_sessions qs ON us.session_id = qs.id
        WHERE us.session_id = %s AND us.user_id = %s
    ''', (session_id, user_id))

    if not info:
        close_db(conn)
        return jsonify({'error': 'Not found'}), 404

    rows = _fetchall(conn, '''
        SELECT
            sec.name       AS section_name,
            sec.order_num  AS sec_order,
            q.order_num    AS q_order,
            q.question_text, q.question_type,
            q.option_a, q.option_b, q.option_c, q.option_d,
            q.correct_answer, q.points,
            ua.selected_answer, ua.is_correct,
            ua.points_earned, ua.reward_code, ua.answered_at
        FROM questions q
        JOIN sections sec ON q.section_id = sec.id
        LEFT JOIN user_answers ua
            ON ua.question_id = q.id
            AND ua.user_session_id = (
                SELECT id FROM user_sessions
                WHERE session_id = %s AND user_id = %s LIMIT 1
            )
        WHERE sec.session_id = %s
        ORDER BY sec.order_num, q.order_num
    ''', (session_id, user_id, session_id))

    opt_cols = {'A': 'option_a', 'B': 'option_b', 'C': 'option_c', 'D': 'option_d'}

    def expand(letters_str, row):
        if not letters_str:
            return ''
        if (row.get('question_type') or 'single') == 'fill_blank':
            return letters_str
        parts = []
        for letter in letters_str.split(','):
            letter = letter.strip().upper()
            col    = opt_cols.get(letter, '')
            text   = row.get(col, '') or ''
            parts.append(f"{letter}: {text}" if text else letter)
        return ' | '.join(parts)

    answers = []
    for r in rows:
        answered = r['selected_answer'] is not None
        answers.append({
            'section':         r['section_name'],
            'question':        r['question_text'],
            'type':            r['question_type'] or 'single',
            'points':          int(r['points'] or 0),
            'points_earned':   float(r['points_earned'] or 0) if answered else None,
            'option_a':        r['option_a'] or '',
            'option_b':        r['option_b'] or '',
            'option_c':        r['option_c'] or '',
            'option_d':        r['option_d'] or '',
            'correct_answer':  expand(r['correct_answer'], r),
            'selected_answer': expand(r['selected_answer'], r) if answered else None,
            'is_correct':      bool(r['is_correct']) if answered else None,
            'reward_code':     r['reward_code'] or '',
            'answered_at':     str(r['answered_at'])[:16] if r['answered_at'] else None,
        })

    close_db(conn)
    return jsonify({
        'name':         info['name'],
        'phone':        info['phone'],
        'session_name': info['session_name'],
        'started_at':   str(info['started_at'])[:16]   if info['started_at']   else None,
        'completed_at': str(info['completed_at'])[:16] if info['completed_at'] else None,
        'flag_count':   int(info['flag_count'] or 0),
        'answers':      answers,
    })


@app.route('/admin/performance/reset', methods=['POST'])
@admin_required
def reset_scores():
    """
    Reset a user's attempt on a session (or ALL users) so they can try again.
    Deletes user_answers and user_sessions rows — as if they never took it.
    POST params:
        session_id : int  (required)
        user_id    : int  (optional — omit to reset ALL users)
    """
    session_id = request.form.get('session_id', type=int)
    user_id    = request.form.get('user_id',    type=int)

    if not session_id:
        flash('No session specified.', 'error')
        return redirect(url_for('admin_performance'))

    conn = get_db()
    try:
        qs_row = _fetchone(conn,
            'SELECT name FROM quiz_sessions WHERE id=%s', (session_id,)
        )

        if not qs_row:
            flash('Session not found.', 'error')
            return redirect(url_for('admin_performance'))

        if user_id:
            # Get all user_session IDs for this user+session
            us_rows = _fetchall(conn,
                'SELECT id FROM user_sessions WHERE session_id=%s AND user_id=%s',
                (session_id, user_id)
            )
            us_ids = [r['id'] for r in us_rows]
            for us_id in us_ids:
                _exec(conn, 'DELETE FROM cheat_flags   WHERE user_session_id=%s', (us_id,))
                _exec(conn, 'DELETE FROM user_answers  WHERE user_session_id=%s', (us_id,))
            _exec(conn,
                'DELETE FROM user_sessions WHERE session_id=%s AND user_id=%s',
                (session_id, user_id)
            )
            user_row = _fetchone(conn,
                'SELECT name FROM users WHERE id=%s', (user_id,)
            )
            name = user_row['name'] if user_row else f'User {user_id}'
            log_action(conn, 'reset_scores_user', entity_type='user',
                       entity_id=user_id, entity_name=name,
                       details=f"Reset scores for {name} on session '{qs_row['name']}'")
            conn.commit()
            flash(f'Reset complete — {name} can now retake "{qs_row["name"]}".', 'success')
        else:
            # Get all user_session IDs for the whole session
            us_rows = _fetchall(conn,
                'SELECT id FROM user_sessions WHERE session_id=%s', (session_id,)
            )
            us_ids = [r['id'] for r in us_rows]
            for us_id in us_ids:
                _exec(conn, 'DELETE FROM cheat_flags   WHERE user_session_id=%s', (us_id,))
                _exec(conn, 'DELETE FROM user_answers  WHERE user_session_id=%s', (us_id,))
            _exec(conn,
                'DELETE FROM user_sessions WHERE session_id=%s', (session_id,)
            )
            log_action(conn, 'reset_scores_all', entity_type='session',
                       entity_id=session_id, entity_name=qs_row['name'],
                       details=f"Reset ALL scores for session '{qs_row['name']}' ({len(us_ids)} attempts deleted)")
            conn.commit()
            flash(f'All scores reset for "{qs_row["name"]}". Everyone can retake it.', 'success')

    finally:
        close_db(conn)

    return redirect(url_for('admin_performance', session_id=session_id))


# ─── Audit Logs ───────────────────────────────────────────────────────────────

@app.route('/admin/audit-logs', methods=['GET', 'POST'])
@admin_required
def admin_audit_logs():
    """View, filter, delete individual logs, mass delete, and purge by age."""
    conn = get_db()

    if request.method == 'POST':
        action = request.form.get('action')

        if action == 'delete_one':
            log_id = request.form.get('log_id', type=int)
            row = _fetchone(conn, 'SELECT action, entity_name FROM audit_logs WHERE id=%s', (log_id,))
            _exec(conn, 'DELETE FROM audit_logs WHERE id=%s', (log_id,))
            log_action(conn, 'delete_audit_log', entity_type='audit_log',
                       entity_id=log_id,
                       details=f"Deleted log #{log_id}: {row['action'] if row else '?'}")
            conn.commit()
            flash('Log entry deleted.', 'success')

        elif action == 'delete_selected':
            ids_raw = request.form.getlist('selected_ids')
            ids = [int(i) for i in ids_raw if i.isdigit()]
            if ids:
                _exec(conn, f'DELETE FROM audit_logs WHERE id = ANY(%s)', (ids,))
                log_action(conn, 'delete_audit_logs_bulk', entity_type='audit_log',
                           details=f"Bulk deleted {len(ids)} log entries: IDs {ids}")
                conn.commit()
                flash(f'{len(ids)} log entries deleted.', 'success')
            else:
                flash('No entries selected.', 'error')

        elif action == 'purge_by_age':
            days = request.form.get('days', type=int)
            if days and days > 0:
                cur = _exec(conn,
                    'DELETE FROM audit_logs WHERE logged_at < (NOW() AT TIME ZONE \'Africa/Nairobi\') - INTERVAL \'1 day\' * %s',
                    (days,)
                )
                count = cur.rowcount
                log_action(conn, 'purge_audit_logs_by_age', entity_type='audit_log',
                           details=f"Purged {count} logs older than {days} days")
                conn.commit()
                flash(f'Purged {count} log entries older than {days} days.', 'success')
            else:
                flash('Please enter a valid number of days.', 'error')

        elif action == 'purge_all':
            cur = _exec(conn, 'DELETE FROM audit_logs')
            count = cur.rowcount
            log_action(conn, 'purge_audit_logs_all', entity_type='audit_log',
                       details=f"Purged ALL {count} audit log entries")
            conn.commit()
            flash(f'All {count} log entries deleted.', 'success')

        close_db(conn)
        # Preserve filter params on redirect
        args = {k: v for k, v in request.args.items() if v}
        return redirect(url_for('admin_audit_logs', **args))

    # ── Filters from query string ──────────────────────────────────────────────
    page        = max(1, request.args.get('page', 1, type=int))
    per_page    = 50
    offset      = (page - 1) * per_page
    filter_cat  = request.args.get('category', '')
    filter_act  = request.args.get('action_filter', '')
    filter_from = request.args.get('date_from', '')
    filter_to   = request.args.get('date_to', '')
    filter_q    = request.args.get('q', '').strip()

    where_clauses = []
    params = []
    if filter_cat:
        where_clauses.append('category = %s');  params.append(filter_cat)
    if filter_act:
        where_clauses.append('action ILIKE %s'); params.append(f'%{filter_act}%')
    if filter_from:
        where_clauses.append('logged_at >= %s'); params.append(filter_from + ' 00:00:00')
    if filter_to:
        where_clauses.append('logged_at <= %s'); params.append(filter_to + ' 23:59:59')
    if filter_q:
        where_clauses.append(
            '(entity_name ILIKE %s OR details ILIKE %s OR action ILIKE %s OR ip_address ILIKE %s)'
        )
        params += [f'%{filter_q}%'] * 4

    where_sql = ('WHERE ' + ' AND '.join(where_clauses)) if where_clauses else ''

    total_row = _fetchone(conn, f'SELECT COUNT(*) FROM audit_logs {where_sql}', params)
    total     = total_row['count'] if total_row else 0
    total_pages = max(1, (total + per_page - 1) // per_page)

    logs = _fetchall(conn,
        f'SELECT * FROM audit_logs {where_sql} ORDER BY logged_at DESC LIMIT %s OFFSET %s',
        params + [per_page, offset]
    )

    # Distinct categories and action names for filter dropdowns
    categories   = [r['category'] for r in _fetchall(conn, 'SELECT DISTINCT category FROM audit_logs ORDER BY category')]
    action_types = [r['action']   for r in _fetchall(conn, 'SELECT DISTINCT action   FROM audit_logs ORDER BY action')]

    close_db(conn)
    return render_template('admin/audit_logs.html',
        logs=logs, total=total, page=page, per_page=per_page, total_pages=total_pages,
        filter_cat=filter_cat, filter_act=filter_act, filter_from=filter_from,
        filter_to=filter_to, filter_q=filter_q,
        categories=categories, action_types=action_types,
    )


@app.route('/admin/settings', methods=['GET', 'POST'])
@admin_required
def admin_settings():
    conn = get_db()
    if request.method == 'POST':
        new_pw = request.form.get('new_password','').strip()
        if new_pw:
            _exec(conn, "UPDATE app_settings SET value=%s WHERE key='admin_password'", (new_pw,))
            log_action(conn, 'change_password', details='Admin password changed')
            conn.commit()
            flash('Password updated!', 'success')
    close_db(conn)
    return render_template('admin/settings.html')


# ═══════════════════════════════════════════════════════════════════════════════
#  FLASK CLI COMMANDS
#  Usage (from the project folder):
#    flask init-db          — create all tables (safe to re-run, won't overwrite)
#    flask reset-db         — ⚠ DROP all tables then recreate (wipes everything)
#    flask reset-db --yes   — skip the confirmation prompt
#    flask create-admin     — set/change the admin password from the terminal
# ═══════════════════════════════════════════════════════════════════════════════

@app.cli.command('init-db')
def cli_init_db():
    """Create all tables (safe: uses CREATE TABLE IF NOT EXISTS)."""
    click.echo('Initialising database…')
    try:
        init_db()
        click.secho('✓ Database initialised successfully.', fg='green')
    except Exception as e:
        click.secho(f'✗ Error: {e}', fg='red')
        raise SystemExit(1)


@app.cli.command('reset-db')
@click.option('--yes', is_flag=True, default=False,
              help='Skip the confirmation prompt.')
def cli_reset_db(yes):
    """DROP all tables then recreate them. ⚠ Destroys all data."""
    if not yes:
        click.secho(
            '\n⚠  WARNING: This will permanently delete ALL data '
            '(users, quizzes, answers, scores).\n',
            fg='yellow', bold=True
        )
        confirmed = click.confirm('Are you sure you want to reset the database?',
                                   default=False)
        if not confirmed:
            click.echo('Aborted — database was NOT changed.')
            return

    click.echo('Dropping all tables…')
    try:
        conn = get_db()
        cur = conn.cursor()
        # Drop in reverse-dependency order so FK constraints don't block
        drop_order = [
            'cheat_flags',
            'user_answers',
            'user_sessions',
            'questions',
            'sections',
            'quiz_sessions',
            'users',
            'app_settings',
            'audit_logs',
        ]
        for table in drop_order:
            cur.execute(f'DROP TABLE IF EXISTS {table} CASCADE')
            click.echo(f'  dropped {table}')
        conn.commit()
        cur.close()
        close_db(conn)
        click.echo('Recreating tables…')
        init_db()
        click.secho('✓ Database reset complete.', fg='green')
    except Exception as e:
        click.secho(f'✗ Error: {e}', fg='red')
        raise SystemExit(1)


@app.cli.command('create-admin')
def cli_create_admin():
    """Set or update the admin panel password."""
    pw = click.prompt('New admin password', hide_input=True,
                      confirmation_prompt='Confirm password')
    if len(pw) < 6:
        click.secho('✗ Password must be at least 6 characters.', fg='red')
        raise SystemExit(1)
    try:
        conn = get_db()
        _exec(conn,
              "INSERT INTO app_settings (key, value) VALUES ('admin_password', %s) "
              "ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value",
              (pw,))
        conn.commit()
        close_db(conn)
        click.secho('✓ Admin password updated.', fg='green')
    except Exception as e:
        click.secho(f'✗ Error: {e}', fg='red')
        raise SystemExit(1)


if __name__ == '__main__':
    # ── NOTE ON CPANEL / PHUSION PASSENGER ────────────────────────────────────
    # On cPanel shared hosting this block NEVER runs.  Passenger imports the
    # `app` object directly via passenger_wsgi.py and handles all serving
    # itself.  Waitress / WAITRESS_THREADS env vars have no effect there.
    #
    # This block is only used when running locally:
    #   python app.py            → Flask dev server (debug mode if FLASK_DEBUG=1)
    #   FLASK_DEBUG=0 python app.py → Waitress (local production test)
    # ──────────────────────────────────────────────────────────────────────────
    init_db()
    port  = int(os.environ.get('PORT', '5000'))
    debug = os.environ.get('FLASK_DEBUG', '').lower() in ('1', 'true', 'yes')
    if debug:
        app.run(host='0.0.0.0', port=port, debug=True)
    else:
        try:
            from waitress import serve
            threads = int(os.environ.get('WAITRESS_THREADS', '8'))
            print(f"Starting Waitress on 0.0.0.0:{port} with {threads} threads")
            serve(app, host='0.0.0.0', port=port, threads=threads)
        except ImportError:
            app.run(host='0.0.0.0', port=port, debug=False, threaded=True)